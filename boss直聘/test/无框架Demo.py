from lxml import etree
from selenium import webdriver
import time
import json
from openpyxl import Workbook


def getcookie(url, cookie_filename):
    driver = webdriver.Chrome()
    driver.get(url)
    time.sleep(15)
    dictCookies = driver.get_cookies()  # 获得所有cookie信息(返回是字典)
    jsonCookies = json.dumps(dictCookies)  # dumps是将dict转化成str格式
    # 登录完成后,将cookies保存到本地文件
    with open(cookie_filename, "w") as fp:
        fp.write(jsonCookies)
        print('Cookies保存成功！')


# 爬取招聘信息并保存到Excel
def scrape_jobs(url, cookie_filename):
    Datalist = []
    for page in range(1, 11):
        print(f"正在处理第 {page} 页...")
        browser = webdriver.Chrome()
        browser.get(url.format(page=page))

        # 注入cookie
        with open(cookie_filename, "r") as fp:
            jsonCookies = fp.read()
        cookies = json.loads(jsonCookies)


        for cookie in cookies:
            browser.add_cookie(cookie)
        browser.get(url.format(page=page))

        time.sleep(6)
        # 获取网页源代码
        content = browser.page_source

        tree = etree.HTML(content)
        li_list = tree.xpath('//li[@class="job-card-wrapper"]')

        for li in li_list:
            time.sleep(3)
            data = []
            # 职位
            title = li.xpath(".//span[@class='job-name']/text()")[0]
            print(type(title))
            data.append(title)
            # 位置
            address = li.xpath(".//span[@class='job-area']/text()")[0]
            data.append(address)
            # 薪资
            salary = li.xpath(".//span[@class='salary']/text()")[0]  # 格式有点奇怪
            data.append(salary)
            # 经验_学历
            job_lable_list = li.xpath(".//ul[@class='tag-list']//text()")
            experience = job_lable_list[0]
            education = job_lable_list[1]
            data.append(experience)
            data.append(education)
            # 公司名
            company = li.xpath(".//h3[@class='company-name']/a/text()")[0]
            data.append(company)
            # 公司行业
            companyType = li.xpath(".//ul[@class='company-tag-list']/li//text()")[0]
            data.append(companyType)
            # 职位技能
            skill_list = li.xpath(".//div[@class='job-card-footer clearfix']//ul[@class='tag-list']/li/text()")
            skill = "|".join(skill_list)
            data.append(skill)
            # 福利 如有全勤奖补贴等
            try:
                boon = li.xpath(".//div[@class='info-desc']/text()")[0]
            except:
                boon = ""
            data.append(boon)
            Datalist.append(data)
            # 输出每条数据
            print(
                f"职位：{title}，公司：{company}，薪资：{salary}，地址：{address}，经验：{experience}，学历：{education}，技能：{skill}，福利：{boon}")

        browser.quit()  # 关闭当前页的浏览器

    save_path = "招聘信息_前端工程师.xls"

    # 列名
    col = ("职位", "位置", "薪资", "经验要求", "学历要求", "公司名", "公司行业", "技能要求", "福利待遇")

    # 创建一个workbook对象
    wb = Workbook()
    # 激活默认的工作表
    ws = wb.active
    # 写入表头
    for i in range(len(col)):
        ws.cell(row=1, column=i + 1, value=col[i])
    # 写入数据
    for i in range(len(Datalist)):
        data = Datalist[i]
        for j in range(len(data)):
            ws.cell(row=i + 2, column=j + 1, value=data[j])

    # 保存工作薄
    wb.save(save_path)
    print(f"数据已保存到 {save_path}")


# 测试运行
url = 'https://www.zhipin.com/web/geek/job?query=%E5%89%8D%E7%AB%AF%E5%B7%A5%E7%A8%8B%E5%B8%88&city=100010000&page={page}'
cookie_filename = r"boss直聘.json"

# 获取Cookie
getcookie('https://www.zhipin.com/web/geek/job-recommend', cookie_filename)

# 爬取招聘信息并保存到Excel
scrape_jobs(url, cookie_filename)
